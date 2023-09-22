
from CTkMessagebox import CTkMessagebox
import customtkinter
from tkinter import filedialog,messagebox
from scipy.signal import convolve2d
from scipy.optimize import curve_fit
from scipy.stats import mode
from sklearn.metrics import r2_score
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from tifffile import imwrite, tifffile
import os
import ntpath
import tensorflow as tf
import imageio
import pandas as pd
from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
from scipy.optimize import minimize
from scipy.special import factorial
from sdtfile import SdtFile

def deconvlucy(image, psf, iterations=10, dampar=0.01, regpar=0.0001):
    # convert the inputs to arrays if they are not already
    image = np.asarray(image)
    psf = np.asarray(psf)

    # compute the Fourier transform of the point spread function
    H = np.fft.fft2(psf, s=image.shape)
    Hconj = np.conj(H)

    # initialize the estimate of the restored image
    x_est = np.ones(image.shape) / np.prod(image.shape)

    for i in range(iterations):
        # compute the estimate of the image using the current estimate and the point spread function
        x_est = x_est * convolve2d(image / convolve2d(x_est, psf, mode='same'), psf[::-1, ::-1], mode='same')

        # compute the Fourier transform of the estimated image
        X_est = np.fft.fft2(x_est)

        # compute the update factor for the estimate of the restored image
        factor = Hconj * H / (Hconj * H + regpar * np.abs(H) ** 2 + dampar * np.abs(X_est) ** 2)

        # compute the updated estimate of the restored image
        x_est = np.real(np.fft.ifft2(factor * np.fft.fft2(image)))

    return x_est


def Convert_ASC_Image(imageDirectory):
    PhotonRawDecayImage = np.loadtxt(imageDirectory, delimiter=' ', skiprows=11)
    photonDecayImage3D = np.reshape(PhotonRawDecayImage, [256, 256, 256])
#     photonDecayImage3D = np.transpose(photonDecayImage3D, [1, 0, 2])
    return photonDecayImage3D
def openFile():
    global tpsfImage
    global filePath
    global intensity
    global tpsfImageRT
    global imgName
    global irf
    filePath =filedialog.askopenfilename(initialdir="C:\\Users\\hulinghao\\PycharmProjects\\Test\\Flim Fit GUI")
    imgName = ntpath.basename(filePath)
    if filePath:
        print("Selected TPSF file:", filePath)
        file_name, file_extension = os.path.splitext(filePath)
        if file_extension == '.asc':
            tpsfImage = Convert_ASC_Image(filePath)
        elif file_extension == '.sdt':
            sdt = SdtFile('test_2023_07_28_11_19.sdt')
            tpsfImage = np.array(sdt.data[0])

        tpsfImage[:, 248:256, :] = 0
        maxIndex = np.argmax(tpsfImage, axis=2)
        globalMaxIndex = mode(maxIndex[maxIndex != 0])[0]
        averageDecay = np.mean(tpsfImage, axis=(0, 1))

        fig, ax = plt.subplots(2,1, figsize=(5, 6),gridspec_kw={'height_ratios': [5, 1]},facecolor='gray')
        im = ax[0].imshow(tpsfImage[:,:,globalMaxIndex], cmap='jet')
        ax[0].axis('off')
        colorbar(im,0,300)
        ax[1].plot(averageDecay)
        try:
            irf
            ax[1].plot(irf * np.max(averageDecay), label="IRF")
            ax[1].legend()
        except NameError:
            print("Please read IRF.")

        canvas = FigureCanvasTkAgg(fig, master=app.Image_frame)
        canvas.draw()
        canvas.get_tk_widget().place(relx=0.025, rely=0.025)

        tpsfImageRT = tpsfImage
    else:
        CTkMessagebox(title="Error", message="No file selected. Please select a TPSF file!!!", icon="cancel")
        print("File not found. Please select a TPSF file.")

def readIRF():
    global irf
    global tpsfImage

    filePath = filedialog.askopenfilename(initialdir="C:\\Linghao Hu\\Project\\FLIM_Fitting\\Result\\Cell Test\\")
    if filePath:
        print("Selected IRF file:", filePath)
        f = open(filePath, "r")
        g_colour_list = []
        for line in f:
            g_colour_list.append(int(line.strip('\n')))
        # print (g_colour_list)
        irf = g_colour_list
        irf = np.asarray(irf)
        irf = irf / max(irf)
        irf = irf.reshape((256, 1))

        maxIndex = np.argmax(tpsfImage, axis=2)
        globalMaxIndex = mode(maxIndex[maxIndex != 0])[0]
        averageDecay = np.mean(tpsfImage, axis=(0, 1))

        fig, ax = plt.subplots(2, 1, figsize=(5, 6), gridspec_kw={'height_ratios': [5, 1]}, facecolor='gray')
        im = ax[0].imshow(tpsfImage[:, :, globalMaxIndex], cmap='jet')
        ax[0].axis('off')
        colorbar(im, 0, 300)
        ax[1].plot(averageDecay)
        ax[1].plot(irf*np.max(averageDecay), label = "IRF")
        ax[1].legend()

        canvas = FigureCanvasTkAgg(fig, master=app.Image_frame)
        canvas.draw()
        canvas.get_tk_widget().place(relx=0.025, rely=0.025)

    else:
        CTkMessagebox(title="Error", message="No file selected. Please select a IRF file!!!", icon="cancel")
        print("File not found. Please select an IRF file.")


def readCellMask():
    global maskImage

    maskFilePath = filedialog.askopenfilename(initialdir="C:\\Linghao Hu\\Project\\FLIM_Fitting\\Result\\Cell Test\\")
    if maskFilePath:
        print("Selected CellMask file:", maskFilePath)
        maskImage = imageio.imread(maskFilePath)
    else:
        messagebox.showerror("Error", "No file selected. Please select a Mask .tif file.")
        print("File not found. Please select a Mask .tif file.")



def Image_Restoration(tpsfImage):
    global tpsfImageRT

    # tpsfImageRT.fill(0)
    imageDiv = Divide_Image_3D(tpsfImage)
    imageDiv = imageDiv.reshape(64, 32, 32, 256, 1)
    imageDiv = (imageDiv - 127.5) / 127.5
    print('The TPSF image is under restoration')
    genImage = FLIM_GAN_Model.predict(imageDiv,batch_size=1)
    genImage = genImage * 127.5 +127.5
    print ('Restoration finish')
    tpsfImageRT = Combine_Image_3D(genImage)

    maxIndex = np.argmax(tpsfImageRT, axis=2)
    globalMaxIndex = mode(maxIndex[maxIndex != 0])[0]
    averageDecay = np.mean(tpsfImageRT, axis=(0, 1))

    fig, ax = plt.subplots(2, 1, figsize=(5, 6),gridspec_kw={'height_ratios': [5, 1]},facecolor='gray')
    im = ax[0].imshow(tpsfImageRT[:, :, globalMaxIndex], cmap='jet')
    ax[0].axis('off')
    colorbar(im,0,300)
    ax[1].plot(averageDecay)

    canvas = FigureCanvasTkAgg(fig, master=app.Image_frame)
    canvas.draw()
    canvas.get_tk_widget().place(relx=0.025, rely=0.025)



def Divide_Image_3D(originalImage):
    width = originalImage.shape[1]
    length = originalImage.shape[2]
    divideNumber = round(width/32)
    xPosition = 0
    yPosition = 0
    counter = 0
    imageNum = (divideNumber ** 2)
    imageDiv = np.ndarray((imageNum, int(32), int(32), int(256)), dtype=np.float32)
    for i in range(imageNum):
        imageDiv[i,:,:,:] = originalImage[xPosition: xPosition + 32, yPosition: yPosition + 32,:]
        xPosition = xPosition + 32
        counter = counter + 1
        if counter == 8:
            xPosition = 0
            yPosition = yPosition + 32
            counter = 0
    return imageDiv

def Combine_Image_3D(imageDiv):
    imageDiv = imageDiv.reshape(64,32,32,256)
    width = 256
    length = 256
    xPosition = 0
    yPosition = 0
    counter = 0
    tpsfImage = np.zeros((256,256,256))
    for i in range(64):
        tpsfImage[xPosition: xPosition + 32, yPosition: yPosition + 32,:] = imageDiv[i,:,:,:]
        xPosition = xPosition + 32
        counter = counter + 1
        if counter == 8:
            xPosition = 0
            yPosition = yPosition + 32
            counter = 0
    return tpsfImage



def biexp_decay(x, a1, tau1, a2, tau2):
    return a1 * np.exp(-x / tau1) + a2 * np.exp(-x / tau2)


### Normalize functoin
def biexp_decay_Norm(x, a1, tau1, tau2):
    return a1 * np.exp(-x / tau1) + (1 - a1) * np.exp(-x / tau2)
def single_decay(x,a,tau):
    return a * np.exp(-x / tau)

def single_decay_Norm(x,tau):
    return np.exp(-x / tau)

# def jacobian(x, a1, tau1, a2, tau2, offset):
#     da1 = np.exp(-x / tau1)
#     dtau1 = -a1 * (x / tau1 ** 2) * np.exp(-x / tau1)
#     da2 = np.exp(-x / tau2)
#     dtau2 = -a2 * (x / tau2 ** 2) * np.exp(-x / tau2)
#     doffset = np.ones_like(x)
#     return np.column_stack((da1, dtau1, da2, dtau2, doffset))
def Get_Threshold():
    global threshold
    threshold = int(app.inputThreshold.get())

def Get_Binning():
    global binningSize
    binningSize = int(app.inputBinningSize.get())

def Binning_Image_Way(choice):
    global binningWay
    global maskImage
    binningLabel = choice
    if binningLabel == 'None':
        binningWay = 0
    elif binningLabel == 'Column':
        binningWay = 1
    elif binningLabel == 'Row':
        binningWay = 2
    elif binningLabel == 'Square':
        binningWay = 3
    else:
        binningWay = 4
        try:
            value = eval('maskImage')
            return True
        except NameError:
            messagebox.showerror("Error", "No Mask file selected. Please select a Mask .tif file if you want to bin in cell.")
            print("File not found. Please select a Mask .tif file.")
            return False
    print("Bin Style:",binningLabel)

### let the binning option to select automatically
def Default_Action_Binning_Way():
    binngingVariable.set(binningChoices[4])
    Binning_Image_Way(None)


def Binning_Image(tpsfImageRT, binningSize, binningWay):
    print("Bin Style:", binningWay)
    print("Bin Size:", binningSize)
    global binningImage
    global maskImage
    binningImage = np.zeros((256, 256, 256))

    if binningWay == 0:
        binningImage = tpsfImageRT
    elif binningWay == 1:
        kernel = np.ones(binningSize)
        for i in range(256):
            timeSliceImage = tpsfImageRT[:,:,i]
            timeSliceImageR = timeSliceImage.T
            binningSliceImageR = np.apply_along_axis(lambda column: np.convolve(column, kernel, mode='same'), axis=1,
                                              arr=timeSliceImageR)
            binningSliceImage = binningSliceImageR.T
            binningImage[:, :, i] = binningSliceImage

    elif binningWay == 2:
        kernel = np.ones(binningSize)
        for i in range(256):
            timeSliceImage = tpsfImageRT[:,:,i]
            binningSliceImage = np.apply_along_axis(lambda row: np.convolve(row, kernel, mode='same'), axis=1,
                                               arr=timeSliceImage)
            binningImage[:,:,i] = binningSliceImage
    elif binningWay == 3:
        kernel = np.ones((binningSize,binningSize))
        for i in range(256):
            timeSliceImage = tpsfImageRT[:, :, i]
            binningSliceImage = convolve2d(timeSliceImage, kernel, mode='same')
            binningImage[:, :, i] = binningSliceImage
    else:
        binningImage = Binning_Cell(tpsfImageRT, maskImage, binningSize)

    maxIndex = np.argmax(binningImage, axis=2)
    globalMaxIndex = mode(maxIndex[maxIndex != 0])[0]
    averageDecay = np.mean(binningImage, axis=(0, 1))

    fig, ax = plt.subplots(2, 1, figsize=(5, 6),gridspec_kw={'height_ratios': [5, 1]},facecolor='gray')
    im = ax[0].imshow(binningImage[:, :, globalMaxIndex], cmap='jet')
    ax[0].axis('off')
    colorbar(im, 0, 300)
    ax[1].plot(averageDecay)

    canvas = FigureCanvasTkAgg(fig, master=app.Image_frame)
    canvas.draw()
    canvas.get_tk_widget().place(relx=0.025, rely=0.025)


def Binning_Cell(tpsfImageT, maskImage, binningSize):

    cellNumber = np.unique(maskImage)
    binCellImage = np.zeros((256, 256, 256))
    kernel = np.ones((binningSize, binningSize))

    for i in range(1, len(cellNumber)):
        maskCopy = np.copy(maskImage)
        maskCopy[maskCopy != i] = 0
        for k in range(256):
            timeSliceImage = tpsfImageT[:, :, k]
            cellTimeSliceImage = timeSliceImage * maskCopy / i
            binningSliceImage = convolve2d(cellTimeSliceImage, kernel, mode='same')
            binningImage[:, :, k] = binningSliceImage

        binCellImage = binCellImage + binningImage

    return binCellImage



def Fit_Component(event):
    global fitComponent
    fittingLabel = componentVariable.get()
    if fittingLabel == 'Single':
        fitComponent = 1
        CTkMessagebox(title="Info", message="Single-component decay fitting only output Tm and r-squared images")
        dispImageOptions = ['tm', 'r2']
        wDisOp.configure(command=Display_Image, values=dispImageOptions,)
    else:
        fitComponent = 2
        dispImageOptions = ['a1','a2','t1','t2','tm','r2']
        wDisOp.configure(command=Display_Image, values=dispImageOptions, )
    print("Fit Component:", fitComponent)

### let the fitting component option to select automatically
def Default_Action_Fitting_Component():
    componentVariable.set(componentChoices[0])
    Fit_Component(None)

def Fit_Method(event):
    global fitMethod
    fittingMethod = fitMethodsVariable.get()
    if fittingMethod == 'MLE':
        fitMethod = 1
    elif fittingMethod == 'Lsq':
        fitMethod = 2
    else:
        fitMethod = 3
    print("Fit Method:", fittingMethod)

def FLIM_Image_Analysis_Single_LSQ(binningImage, irf, threshold):
    global rSquareImage
    global a1Image
    global tmImage

    print("Threshold:", threshold, "Fitting Start...")
    timePeroid = np.arange(1, 256) * 12.5 / 256
    a1Image = np.zeros((256, 256))
    tmImage = np.zeros((256, 256))
    rSquareImage = np.zeros((256, 256))

    tpsfImageMax = np.amax(binningImage, axis=2)
    indices = np.where(tpsfImageMax > threshold)

    # maxIndex = np.argmax(tpsfImage, axis=2)
    # globalMaxIndex = mode(maxIndex[maxIndex != 0]).mode[0]

    n = len(indices[0])
    iter_step = 1 / n
    progress_step = iter_step
    app.fitProgress.start()

    for k in range(n):
        i = indices[0][k]
        j = indices[1][k]
        predictDecay = binningImage[i, j, :]
        predictDecay = predictDecay.reshape((256, 1))
        deconvDecay = deconvlucy(predictDecay, irf[49:79], iterations=10, dampar=0, regpar=5)

        nan_indices = np.isnan(deconvDecay)
        if (nan_indices.any()):
            a1Image[i, j] = 0
            tmImage[i, j] = 0
            rSquareImage[i, j] = 0

        else:
            locMaxIndex = np.argmax(deconvDecay)
            peakIndex = locMaxIndex

            #         peakIndex = globalMaxIndex
            y = deconvDecay[peakIndex:256]
            y = np.squeeze(y)
            y = np.trim_zeros(y, 'b')

            x = timePeroid[0:np.size(y)]

            popt, pcov = curve_fit(single_decay, x, y, bounds=((0, 0), (30, 5)),ftol=0.1, xtol=0.1)
            model_predictions = single_decay(x, *popt)
            r_squared = r2_score(y, model_predictions)

            a = popt[0]
            t = popt[1]

            a1Image[i, j] = a
            tmImage[i, j] = t
            rSquareImage[i, j] = r_squared

            app.fitProgress.set(progress_step)
            progress_step += iter_step
            app.fitProgress.update_idletasks()
            app.fitProgress.stop()
            app.update()

    a1Image = np.nan_to_num(a1Image, nan=0.0)
    tmImage = np.nan_to_num(tmImage, nan=0.0)
    rSquareImage = np.nan_to_num(rSquareImage, nan=0.0)

    fig, ax = plt.subplots(2, 1, figsize=(5, 6), gridspec_kw={'height_ratios': [5, 1]}, facecolor='gray')
    im = ax[0].imshow(tmImage, cmap='jet')
    ax[0].axis('off')
    colorbar(im, 0.5, 1.5)
    tmArray = tmImage.flatten()
    tmArray = tmArray[tmArray != 0]

    hist, edges = np.histogram(tmArray, range=(0.5, 1.5), bins=50)
    freq = hist / float(hist.sum())
    width = np.diff(edges)  # edges is bins

    ax[1].bar(edges[1:], freq, width=width, align="edge", ec="k")
    ax[1].set(ylabel='frequency')
    ax[1].set_title('Tm')

    canvas = FigureCanvasTkAgg(fig, master=app.Image_frame)
    canvas.draw()
    canvas.get_tk_widget().place(relx=0.025, rely=0.025)

    print("Fitting Done")
    return a1Image, tmImage, rSquareImage
def FLIM_Image_Analysis_Single_LSQ_Norm(binningImage, irf, threshold):
    global rSquareImage
    global a1Image
    global tmImage

    print("Threshold:", threshold, "Fitting Start...")
    timePeroid = np.arange(1, 256) * 12.5 / 256
    a1Image = np.zeros((256, 256))
    tmImage = np.zeros((256, 256))
    rSquareImage = np.zeros((256, 256))

    tpsfImageMax = np.amax(binningImage, axis=2)
    indices = np.where(tpsfImageMax > threshold)

    # maxIndex = np.argmax(tpsfImage, axis=2)
    # globalMaxIndex = mode(maxIndex[maxIndex != 0]).mode[0]

    n = len(indices[0])
    iter_step = 1 / n
    progress_step = iter_step
    app.fitProgress.start()

    for k in range(n):
        i = indices[0][k]
        j = indices[1][k]
        predictDecay = binningImage[i, j, :]
        predictDecay = predictDecay.reshape((256, 1))
        deconvDecay = deconvlucy(predictDecay, irf[49:79], iterations=10, dampar=0, regpar=5)

        nan_indices = np.isnan(deconvDecay)
        if (nan_indices.any()):
            a1Image[i, j] = 0
            tmImage[i, j] = 0
            rSquareImage[i, j] = 0

        else:
            locMaxIndex = np.argmax(deconvDecay)
            peakIndex = locMaxIndex

            #         peakIndex = globalMaxIndex
            y = deconvDecay[peakIndex:256]
            y = np.squeeze(y)
            y = np.trim_zeros(y, 'b')

            y = y.astype(int)
            yNorm = y / np.max(y)
            x = timePeroid[0:np.size(yNorm)]


            popt, pcov = curve_fit(single_decay_Norm, x, yNorm,ftol=0.1, xtol=0.1)
            model_predictions = single_decay_Norm(x, *popt)
            r_squared = r2_score(yNorm, model_predictions)

            a = 1
            t = popt[0]

            a1Image[i, j] = a
            tmImage[i, j] = t
            rSquareImage[i, j] = r_squared

            app.fitProgress.set(progress_step)
            progress_step += iter_step
            app.fitProgress.update_idletasks()
            app.fitProgress.stop()
            app.update()

    a1Image = np.nan_to_num(a1Image, nan=0.0)
    tmImage = np.nan_to_num(tmImage, nan=0.0)
    rSquareImage = np.nan_to_num(rSquareImage, nan=0.0)

    fig, ax = plt.subplots(2, 1, figsize=(5, 6), gridspec_kw={'height_ratios': [5, 1]}, facecolor='gray')
    im = ax[0].imshow(tmImage, cmap='jet')
    ax[0].axis('off')
    colorbar(im, 0.5, 1.5)
    tmArray = tmImage.flatten()
    tmArray = tmArray[tmArray != 0]

    hist, edges = np.histogram(tmArray, range=(0.5, 1.5), bins=50)
    freq = hist / float(hist.sum())
    width = np.diff(edges)  # edges is bins

    ax[1].bar(edges[1:], freq, width=width, align="edge", ec="k")
    ax[1].set(ylabel='frequency')
    ax[1].set_title('Tm')

    canvas = FigureCanvasTkAgg(fig, master=app.Image_frame)
    canvas.draw()
    canvas.get_tk_widget().place(relx=0.025, rely=0.025)

    print("Fitting Done")
    return a1Image, tmImage, rSquareImage
def FLIM_Image_Analysis_Single_MLE(binningImage, irf, threshold):
    global a1Image
    global tmImage
    global rSquareImage

    print("Threshold:", threshold, "Fitting Start...")
    timePeroid = np.arange(1, 256) * 12.5 / 256
    a1Image = np.zeros((256, 256))
    tmImage = np.zeros((256, 256))
    rSquareImage = np.zeros((256, 256))

    tpsfImageMax = np.amax(binningImage, axis=2)
    indices = np.where(tpsfImageMax > threshold)

    n = len(indices[0])
    iter_step = 1 / n
    progress_step = iter_step
    app.fitProgress.start()

    for k in range(n):
        i = indices[0][k]
        j = indices[1][k]
        predictDecay = binningImage[i, j, :]
        predictDecay = predictDecay.reshape((256, 1))
        deconvDecay = deconvlucy(predictDecay, irf[49:79], iterations=10, dampar=0, regpar=5)
        shiftDeconvDecay = np.roll(deconvDecay, 7)
        locMaxIndex = np.argmax(shiftDeconvDecay)

        peakIndex = locMaxIndex
        #         peakIndex = globalMaxIndex
        y = shiftDeconvDecay[peakIndex:256]
        y = np.squeeze(y)
        y = y.astype(int)
        x = timePeroid[0:np.size(y)]
        yReal = y

        def MLE(params):
            """ find the max likelihood """
            a, tau= params
            yPred = single_decay(x, a, tau)
            poss = poisson(yPred, yReal)
            sumPoss = -np.sum(poss)
            return sumPoss

        guess = np.array([20, 1.0])
        bnds = ((20, 50), (1, 3))

        result = minimize(MLE, x0=guess, bounds=bnds, method='TNC', tol=0.5)

        A, T,= result.x
        y_fitted = A * np.exp(-x / T)
        r_squared = r2_score(yReal, y_fitted)

        a1 = A
        tm = T

        a1Image[i, j] = a1
        tmImage[i, j] = tm
        rSquareImage[i, j] = r_squared

        app.fitProgress.set(progress_step)
        progress_step += iter_step
        app.fitProgress.update_idletasks()
        app.fitProgress.stop()
        app.update()

    a1Image = np.nan_to_num(a1Image, nan=0.0)
    tmImage = np.nan_to_num(tmImage, nan=0.0)
    rSquareImage = np.nan_to_num(rSquareImage, nan=0.0)

    fig, ax = plt.subplots(2, 1, figsize=(5, 6), gridspec_kw={'height_ratios': [5, 1]}, facecolor='gray')
    im = ax[0].imshow(tmImage, cmap='jet')
    ax[0].axis('off')
    colorbar(im, 0.5, 1.5)
    tmArray = tmImage.flatten()
    tmArray = tmArray[tmArray != 0]

    hist, edges = np.histogram(tmArray, range=(0.5, 1.5), bins=50)
    freq = hist / float(hist.sum())
    width = np.diff(edges)  # edges is bins

    # bins = np.linspace(0.5, 1.5, 40)  # Customize the bin edges as needed
    # ax[1].hist(tmArray, 40, density=True, edgecolor='black')

    ax[1].bar(edges[1:], freq, width=width, align="edge", ec="k")
    ax[1].set(ylabel='frequency')
    ax[1].set_title('Tm')

    canvas = FigureCanvasTkAgg(fig, master=app.Image_frame)
    canvas.draw()
    canvas.get_tk_widget().place(relx=0.025, rely=0.025)

    print("Fitting Done")
    return a1Image, tmImage, rSquareImage
def FLIM_Image_Analysis_Bi_LSQ_Norm(binningImage, irf, threshold):
    global a1Image
    global a2Image
    global t1Image
    global t2Image
    global tmImage
    global rSquareImage

    print("Threshold:", threshold, "Fitting Start...")
    timePeroid = np.arange(0, 255) * 12.5 / 256
    a1Image = np.zeros((256, 256))
    a2Image = np.zeros((256, 256))
    t1Image = np.zeros((256, 256))
    t2Image = np.zeros((256, 256))
    tmImage = np.zeros((256, 256))
    rSquareImage = np.zeros((256, 256))

    tpsfImageMax = np.amax(binningImage, axis=2)
    indices = np.where(tpsfImageMax > threshold)

    maxIndex = np.argmax(tpsfImage, axis=2)
    globalMaxIndex = mode(maxIndex[maxIndex != 0])[0]
    n = len(indices[0])
    iter_step = 1 / n
    progress_step = iter_step
    app.fitProgress.start()

    for k in range(n):
        i = indices[0][k]
        j = indices[1][k]
        predictDecay = binningImage[i, j, :]
        predictDecay = predictDecay.reshape((256, 1))
        deconvDecay = deconvlucy(predictDecay, irf[49:79], iterations=10, dampar=0, regpar=5)

        shiftDeconvDecay = np.roll(deconvDecay, 7)
        locMaxIndex = np.argmax(shiftDeconvDecay)
        # peakIndex = globalMaxIndex
        peakIndex = locMaxIndex

        y = deconvDecay[peakIndex:256]
        y = np.squeeze(y)

        # y = np.trim_zeros(y, 'b')
        # first_negative_index = np.argmax(y < 0)
        # # Cut the array from the first negative index onwards
        # y = y[:first_negative_index]

        y = y.astype(int)
        yNorm = y / np.max(y)
        x = timePeroid[0:np.size(yNorm)]


        ### LSQ Curve Normalization before fit
        popt, pcov = curve_fit(biexp_decay_Norm, x, yNorm, bounds=((0.5, 0.3, 1), (1, 0.32, 3)), ftol=0.5, xtol=0.5)

        model_predictions = biexp_decay_Norm(x, *popt)
        r_squared = r2_score(yNorm, model_predictions)


        a1 = popt[0]
        a2 = 1 - a1
        t1 = popt[1]
        t2 = popt[2]
        tm = a1 * t1 + a2 * t2

        a1Image[i, j] = a1
        a2Image[i, j] = a2
        t1Image[i, j] = t1
        t2Image[i, j] = t2
        tmImage[i, j] = tm
        rSquareImage[i, j] = r_squared
        # Fit Progress
        app.fitProgress.set(progress_step)
        progress_step += iter_step
        app.fitProgress.update_idletasks()
        app.fitProgress.stop()
        app.update()


    a1Image = np.nan_to_num(a1Image, nan=0.0)
    a2Image = np.nan_to_num(a2Image, nan=0.0)
    t1Image = np.nan_to_num(t1Image, nan=0.0)
    t2Image = np.nan_to_num(t2Image, nan=0.0)
    tmImage = np.nan_to_num(tmImage, nan=0.0)
    rSquareImage = np.nan_to_num(rSquareImage, nan=0.0)

    fig, ax = plt.subplots(2, 1, figsize=(5, 6), gridspec_kw={'height_ratios': [5, 1]}, facecolor='gray')
    im = ax[0].imshow(tmImage, cmap='jet')
    ax[0].axis('off')
    colorbar(im, 0.5, 1.5)
    tmArray = tmImage.flatten()
    tmArray = tmArray[tmArray != 0]

    hist, edges = np.histogram(tmArray, range=(0.5, 1.5), bins=50)
    freq = hist / float(hist.sum())
    width = np.diff(edges)  # edges is bins


    ax[1].bar(edges[1:], freq, width=width, align="edge", ec="k")
    ax[1].set(ylabel='frequency')
    ax[1].set_title('Tm')

    canvas = FigureCanvasTkAgg(fig, master=app.Image_frame)
    canvas.draw()
    canvas.get_tk_widget().place(relx=0.025, rely=0.025)

    print("Fitting Done")
    return a1Image, a2Image, t1Image, t2Image, tmImage, rSquareImage

def FLIM_Image_Analysis_Bi_LSQ(binningImage, irf, threshold):
    global a1Image
    global a2Image
    global t1Image
    global t2Image
    global tmImage
    global rSquareImage

    print("Threshold:", threshold, "Fitting Start...")
    timePeroid = np.arange(0, 255) * 12.5 / 256
    a1Image = np.zeros((256, 256))
    a2Image = np.zeros((256, 256))
    t1Image = np.zeros((256, 256))
    t2Image = np.zeros((256, 256))
    tmImage = np.zeros((256, 256))
    rSquareImage = np.zeros((256, 256))

    tpsfImageMax = np.amax(binningImage, axis=2)
    indices = np.where(tpsfImageMax > threshold)

    maxIndex = np.argmax(tpsfImage, axis=2)
    globalMaxIndex = mode(maxIndex[maxIndex != 0])[0]
    n = len(indices[0])
    iter_step = 1 / n
    progress_step = iter_step
    app.fitProgress.start()

    for k in range(n):
        i = indices[0][k]
        j = indices[1][k]

        predictDecay = tpsfImage[i, j, :]
        predictDecay = predictDecay.reshape((256, 1))
        deconvDecay = deconvlucy(predictDecay, irf[49:79], iterations=10, dampar=0, regpar=5)
        shiftDeconvDecay = np.roll(deconvDecay, 7)

        #         peakIndex = globalMaxIndex
        locMaxIndex = np.argmax(shiftDeconvDecay)
        peakIndex = locMaxIndex

        y = shiftDeconvDecay[peakIndex:256]
        y = np.squeeze(y)
        y = y.astype(int)
        x = timePeroid[0:np.size(y)]

        bnds = ((20, 0.3, 0, 1), (50, 0.32, 20, 5))
        popt, pcov = curve_fit(biexp_decay, x, y, bounds=bnds, ftol=0.5, xtol=0.5)
        y_fitted = biexp_decay(x, *popt)
        r_squared = r2_score(y, y_fitted)

        a1 = popt[0] / (popt[0] + popt[2])
        a2 = 1 - a1
        t1 = popt[1]
        t2 = popt[3]
        tm = a1 * t1 + a2 * t2

        a1Image[i, j] = a1
        a2Image[i, j] = a2
        t1Image[i, j] = t1
        t2Image[i, j] = t2
        tmImage[i, j] = tm
        rSquareImage[i, j] = r_squared

        app.fitProgress.set(progress_step)
        progress_step += iter_step
        app.fitProgress.update_idletasks()
        app.fitProgress.stop()
        app.update()

    a1Image = np.nan_to_num(a1Image, nan=0.0)
    a2Image = np.nan_to_num(a2Image, nan=0.0)
    t1Image = np.nan_to_num(t1Image, nan=0.0)
    t2Image = np.nan_to_num(t2Image, nan=0.0)
    tmImage = np.nan_to_num(tmImage, nan=0.0)
    rSquareImage = np.nan_to_num(rSquareImage, nan=0.0)

    fig, ax = plt.subplots(2, 1, figsize=(5, 6), gridspec_kw={'height_ratios': [5, 1]}, facecolor='gray')
    im = ax[0].imshow(tmImage, cmap='jet')
    ax[0].axis('off')
    colorbar(im, 0.5, 1.5)
    tmArray = tmImage.flatten()
    tmArray = tmArray[tmArray != 0]

    hist, edges = np.histogram(tmArray, range=(0.5, 1.5), bins=50)
    freq = hist / float(hist.sum())
    width = np.diff(edges)  # edges is bins


    ax[1].bar(edges[1:], freq, width=width, align="edge", ec="k")
    ax[1].set(ylabel='frequency')
    ax[1].set_title('Tm')

    canvas = FigureCanvasTkAgg(fig, master=app.Image_frame)
    canvas.draw()
    canvas.get_tk_widget().place(relx=0.025, rely=0.025)

    print("Fitting Done")
    return a1Image, a2Image, t1Image, t2Image, tmImage, rSquareImage
def poisson(k, lamb):
    """poisson pdf, parameter lamb is the fit parameter"""
    return (lamb**k/factorial(k)) * np.exp(-lamb)
def FLIM_Image_Analysis_Bi_MLE(binningImage, irf, threshold):
    global a1Image
    global a2Image
    global t1Image
    global t2Image
    global tmImage
    global rSquareImage

    print("Threshold:", threshold, "Fitting Start...")
    timePeroid = np.arange(1, 256) * 12.5 / 256
    a1Image = np.zeros((256, 256))
    a2Image = np.zeros((256, 256))
    t1Image = np.zeros((256, 256))
    t2Image = np.zeros((256, 256))
    tmImage = np.zeros((256, 256))
    rSquareImage = np.zeros((256, 256))

    tpsfImageMax = np.amax(binningImage, axis=2)
    indices = np.where(tpsfImageMax > threshold)

    # maxIndex = np.argmax(tpsfImage, axis=2)
    # globalMaxIndex = mode(maxIndex[maxIndex != 0]).mode[0]
    n = len(indices[0])
    iter_step = 1 / n
    progress_step = iter_step
    app.fitProgress.start()

    for k in range(n):
        i = indices[0][k]
        j = indices[1][k]
        predictDecay = binningImage[i, j, :]
        predictDecay = predictDecay.reshape((256, 1))
        deconvDecay = deconvlucy(predictDecay, irf[49:79], iterations=10, dampar=0, regpar=5)
        shiftDeconvDecay = np.roll(deconvDecay, 7)
        locMaxIndex = np.argmax(shiftDeconvDecay)

        peakIndex = locMaxIndex
        #         peakIndex = globalMaxIndex
        y = shiftDeconvDecay[peakIndex:256]
        y = np.squeeze(y)
        y = y.astype(int)
        x = timePeroid[0:np.size(y)]
        yReal = y

        def MLE(params):
            """ find the max likelihood """
            a1, tau1, a2, tau2 = params
            yPred = biexp_decay(x, a1, tau1, a2, tau2)
            poss = poisson(yPred, yReal)
            sumPoss = -np.sum(poss)
            return sumPoss

        guess = np.array([20, 0.3, 10, 2])
        bnds = ((20, 50), (0.2, 0.7), (0, 20), (1, 5))

        result = minimize(MLE, x0=guess, bounds=bnds, method='TNC', tol = 0.5)

        A1, T1, A2, T2 = result.x
        y_fitted = A1 * np.exp(-x / T1) + A2 * np.exp(-x / T2)
        r_squared = r2_score(yReal, y_fitted)

        a1 = A1 / (A1 + A2)
        a2 = 1 - a1
        t1 = T1
        t2 = T2
        tm = a1 * t1 + a2 * t2

        a1Image[i, j] = a1
        a2Image[i, j] = a2
        t1Image[i, j] = t1
        t2Image[i, j] = t2
        tmImage[i, j] = tm
        rSquareImage[i, j] = r_squared


        app.fitProgress.set(progress_step)
        progress_step += iter_step
        app.fitProgress.update_idletasks()
        app.fitProgress.stop()
        app.update()

    a1Image = np.nan_to_num(a1Image, nan=0.0)
    a2Image = np.nan_to_num(a2Image, nan=0.0)
    t1Image = np.nan_to_num(t1Image, nan=0.0)
    t2Image = np.nan_to_num(t2Image, nan=0.0)
    tmImage = np.nan_to_num(tmImage, nan=0.0)
    rSquareImage = np.nan_to_num(rSquareImage, nan=0.0)





    fig, ax = plt.subplots(2, 1, figsize=(5, 6), gridspec_kw={'height_ratios': [5, 1]}, facecolor='gray')
    im = ax[0].imshow(tmImage, cmap='jet')
    ax[0].axis('off')
    colorbar(im, 0.5, 1.5)
    tmArray = tmImage.flatten()
    tmArray = tmArray[tmArray != 0]

    hist, edges = np.histogram(tmArray, range = (0.5,1.5),bins = 50)
    freq = hist / float(hist.sum())
    width = np.diff(edges)  # edges is bins



    ax[1].bar(edges[1:], freq, width=width, align="edge", ec="k")
    ax[1].set(ylabel='frequency')
    ax[1].set_title('Tm')

    canvas = FigureCanvasTkAgg(fig, master=app.Image_frame)
    canvas.draw()
    canvas.get_tk_widget().place(relx=0.025, rely=0.025)

    print("Fitting Done")
    return a1Image, a2Image, t1Image, t2Image, tmImage, rSquareImage

def FLIM_Fit():
    if fitComponent == 1:
        if fitMethod == 1:
            FLIM_Image_Analysis_Single_MLE(binningImage, irf, threshold)
        elif fitMethod == 2:
            FLIM_Image_Analysis_Single_LSQ(binningImage, irf, threshold)
        else:
            FLIM_Image_Analysis_Single_LSQ_Norm(binningImage, irf, threshold)
    else:
        if fitMethod == 1:
            FLIM_Image_Analysis_Bi_MLE(binningImage, irf, threshold)
        elif fitMethod == 2:
            FLIM_Image_Analysis_Bi_LSQ(binningImage, irf, threshold)
        else:
            FLIM_Image_Analysis_Bi_LSQ_Norm(binningImage, irf, threshold)


def Display_Image(event):
    displayImage = dispImageVariable.get()
    if displayImage == 'a1':
        Show_Image_Histogram(a1Image, 0.5, 1, displayImage)
    elif displayImage == 'a2':
        Show_Image_Histogram(a2Image, 0, 0.5, displayImage)
    elif displayImage == 't1':
        Show_Image_Histogram(t1Image, 0.2, 0.7, displayImage)
    elif displayImage == 't2':
        Show_Image_Histogram(t2Image, 1, 3, displayImage)
    elif displayImage == 'tm':
        Show_Image_Histogram(tmImage, 0.5, 1.5, displayImage)
    elif displayImage == 'r2':
        Show_Image_Histogram(rSquareImage, 0.5, 1, displayImage)
def Show_Image_Histogram(image, min, max, title):
    fig, ax = plt.subplots(2, 1, figsize=(5, 6),gridspec_kw={'height_ratios': [5, 1]},facecolor='gray')
    im = ax[0].imshow(image, cmap='jet')
    ax[0].axis('off')
    colorbar(im,min,max)
    array = image.flatten()
    array = array[array != 0]
    # bins = np.linspace(min, max, 40)  # Customize the bin edges as needed
    # ax[1].hist(array, 40, density=True, edgecolor='black')
    # ax[1].set_title(title)

    hist, edges = np.histogram(array, range=(min, max), bins=50)
    freq = hist / float(hist.sum())
    width = np.diff(edges)  # edges is bins
    ax[1].bar(edges[1:], freq, width=width, align="edge", ec="k")
    ax[1].set(ylabel='frequency')
    ax[1].set_title(title)

    canvas = FigureCanvasTkAgg(fig, master=app.Image_frame)
    canvas.draw()
    canvas.get_tk_widget().place(relx=0.025, rely=0.025)

def colorbar(mappable,minR, maxR):
    from mpl_toolkits.axes_grid1 import make_axes_locatable
    import matplotlib.pyplot as plt
    last_axes = plt.gca()
    ax = mappable.axes
    fig = ax.figure
    divider = make_axes_locatable(ax)
    cax = divider.append_axes("right", size="5%", pad=0.05)
    mappable.set_clim(vmin=minR, vmax=maxR)
    cbar = fig.colorbar(mappable, cax=cax)
    plt.sca(last_axes)
    return cbar

def Image_Save(fitComponent,filePath,imgName):
    imgName = imgName[:-4]
    filePosition = os.path.dirname(filePath)
    imageName = filePosition + '/' + imgName
    if fitComponent == 1:
        Single_Fit_Save(imageName, a1Image, tmImage, rSquareImage)
    else:
        Bi_Fit_Save(imageName, a1Image, a2Image, t1Image, t2Image, tmImage, rSquareImage)


def Single_Fit_Save(imageName, a1Image, tmImage, rSquareImage):
    imwrite(imageName +"_a.tif", a1Image.astype('float32'))
    imwrite(imageName + "_t.tif", tmImage.astype('float32'))
    imwrite(imageName + "_RS.tif", rSquareImage.astype('float32'))


def Bi_Fit_Save(imageName, a1Image, a2Image, t1Image, t2Image, tmImage, rSquareImage):
    imwrite(imageName + "_a1.tif", a1Image.astype('float32'))
    imwrite(imageName + "_a2.tif", a2Image.astype('float32'))
    imwrite(imageName + "_t1.tif", t1Image.astype('float32'))
    imwrite(imageName + "_t2.tif", t2Image.astype('float32'))
    imwrite(imageName + "_tm.tif", tmImage.astype('float32'))
    imwrite(imageName + "_RS.tif", rSquareImage.astype('float32'))


def Calculate_Life_Cell(a1Image, t1Image, t2Image, tmImage, tpsfImage, maskImage):
    global a1CellList
    global t1CellList
    global t2CellList
    global tmCellList
    global intensityList
    global saveData
    # CTkMessagebox(title="Info", message="Apply mask only works for bi-exponential decay fitting with inputting a cellular mask")

    cellNumber = np.unique(maskImage)
    intensityImage = np.sum(tpsfImage, axis=2)



    a1CellList = np.zeros(len(cellNumber)-1)
    t1CellList = np.zeros(len(cellNumber)-1)
    t2CellList = np.zeros(len(cellNumber)-1)
    tmCellList = np.zeros(len(cellNumber)-1)
    intensityList = np.zeros(len(cellNumber)-1)

    print('Calculating fluorescence lifetime parameters for each cell ...')
    for i in range(1, len(cellNumber)):
        maskCopy = np.copy(maskImage)
        maskCopy[maskCopy != i] = 0
        maskCopy = np.array(maskCopy/i,dtype=bool)
        maskCella1 = a1Image[maskCopy]
        maskCellt1 = t1Image[maskCopy]
        maskCellt2 = t2Image[maskCopy]
        maskCelltm = tmImage[maskCopy]
        maskCellInt = intensityImage[maskCopy]

        a1Cell = np.mean(maskCella1[maskCella1 != 0])
        t1Cell = np.mean(maskCellt1[maskCellt1 != 0])
        t2Cell = np.mean(maskCellt2[maskCellt2 != 0])
        tmCell = np.mean(maskCelltm[maskCelltm != 0])
        intCell = np.mean(maskCellInt[maskCellInt != 0])


        a1CellList[i - 1] = a1Cell
        t1CellList[i - 1] = t1Cell
        t2CellList[i - 1] = t2Cell
        tmCellList[i - 1] = tmCell
        intensityList[i - 1] = intCell

    saveData = pd.DataFrame({'a1': a1CellList, 't1': t1CellList, 't2':t2CellList,'tm': tmCellList,'Intensity': intensityList}, columns=['a1', 't1','t2','tm','Intensity'])

    print('Calculating fluorescence lifetime parameters for each cell is done. ')
    maskCopyShow = np.copy(maskImage)
    maskCopyShow[maskCopyShow != 0] = 1
    tmImageAfterMask = np.multiply(tmImage, maskCopyShow)
    Show_Image_Histogram(tmImageAfterMask, 0.5, 1.5,'Tm (mask)')
# Read Excel function
def Select_Excel_File():
    global saveExcelPath

    saveExcelPath = filedialog.askopenfilename(initialdir="C:\\Linghao Hu\\Project\\FLIM_Fitting\\Result\\Cell Test\\")
    if saveExcelPath:
        print("The excel file to save the lifetime values", saveExcelPath)
        # saveExcelText.delete(1.0, tk.END)  # Clear any previous content
        # saveExcelText.insert(tk.END, saveExcelPath)
    else:
        messagebox.showerror("Error", "No excel file selected. Please select an excel file to save lifetime values.")
        print("Excel file not found. Please select an excel file to save lifetime values.")


def Save_Values_Excel(imgName):
    global saveExcelPath
    global saveData

    sheetName = app.saveSheet.get()

    # Load the Excel file
    book = load_workbook(saveExcelPath)


    # Select the first sheet (you can use book.sheetnames to get the sheet names)
    try:
        book.sheets = dict((ws.title, ws) for ws in book.worksheets)
        sheet = book.sheets[sheetName]
    except KeyError:
        # If the sheet doesn't exist, create a new one
        sheet = book.create_sheet(sheetName)
        book.save(saveExcelPath)

    # Find the last row in the sheet
    last_row = sheet.max_row # Add 1 to append after the last row
    existing_data = pd.read_excel(saveExcelPath, sheet_name = sheetName)
    empty_data = pd.DataFrame({'a1':[imgName], 't1': ['NA'], 't2':['NA'],'tm': ['NA'],'Intensity': ['NA']})
    # Create a pandas ExcelWriter object

    writer = pd.ExcelWriter(saveExcelPath, engine='openpyxl')
    writer.book = book
    print(['Save the lifetime values of each cell to the Excel',sheetName])
    # Write the DataFrame to the Excel file starting from the last row
    if last_row == 1:
        combined_data = pd.concat([empty_data, saveData])
        combined_data.to_excel(writer, index=False, sheet_name=sheetName, header=True)
    else:
        combined_data = pd.concat([existing_data, empty_data, saveData], ignore_index=True)
        combined_data.to_excel(writer, index=False,sheet_name=sheetName, header=True)

    # Save the ExcelWriter to the Excel file
    writer.save()
    book.close()
    print('Save done')

## Create the GUI
customtkinter.set_appearance_mode("System")  # Modes: system (default), light, dark
customtkinter.set_default_color_theme("blue")  # Themes: blue (default), dark-blue, green


app = customtkinter.CTk()  # create CTk window like you do with the Tk window
app.title('FLIM Analysis')
app.geometry("810x700")

app.grid_rowconfigure(1, weight=1)
app.grid_columnconfigure(1, weight=1)

# Creat Navigation frame
app.Navigation_frame = customtkinter.CTkFrame(app, corner_radius=0)
app.Navigation_frame.grid(row=0, column=0, sticky="nsew")

app.Navigation_frame.grid_columnconfigure(0, weight=1)
app.Navigation_frame.grid_rowconfigure(16, weight=1)


### Read Images
app.imageRead_Title = customtkinter.CTkLabel(app.Navigation_frame, text="Image Read", compound="left", font=customtkinter.CTkFont(size=20, weight="bold"))
app.imageRead_Title.grid(row=0, column=0, padx=20, pady=10,columnspan=2)

app.tpsfReadButton = customtkinter.CTkButton(app.Navigation_frame, corner_radius=0, height=40, border_spacing=10,
                                           text="Open TPSF",
                                           fg_color="transparent", text_color=("gray10", "gray90"),
                                           hover_color=("gray70", "gray30"),
                                           anchor="w",command=lambda: openFile())
app.tpsfReadButton.grid(row=1, column=0, sticky="ew")

app.IRFReadButton = customtkinter.CTkButton(app.Navigation_frame, corner_radius=0, height=40, border_spacing=10,
                                              text="Read IRF",
                                              fg_color="transparent", text_color=("gray10", "gray90"),
                                              hover_color=("gray70", "gray30"),
                                              anchor="w",command=lambda: readIRF())
app.IRFReadButton.grid(row=1, column=1, sticky="ew")


app.ReadMask_button = customtkinter.CTkButton(app.Navigation_frame, corner_radius=0, height=40, border_spacing=10,
                                              text="Read Cell Mask",
                                              fg_color="transparent", text_color=("gray10", "gray90"),
                                              hover_color=("gray70", "gray30"),
                                              anchor="w", command=lambda: readCellMask())
app.ReadMask_button.grid(row=2, column=0, sticky="ew",columnspan=2)

app.FlimAnalysis_imageRestorationButton = customtkinter.CTkButton(app.Navigation_frame,text="Image Restoration", command=lambda: Image_Restoration(tpsfImage),
                                                 corner_radius=0, height=40, border_spacing=10,
                                                 fg_color="transparent", text_color=("gray10", "gray90"),
                                                 hover_color=("gray70", "gray30"),
                                                 anchor="w")

app.FlimAnalysis_imageRestorationButton.grid(row=3, column=0,sticky="ew",columnspan=2)

### FLIM Fit
app.imageFit_Title = customtkinter.CTkLabel(app.Navigation_frame, text="Fit FLIM", compound="left", font=customtkinter.CTkFont(size=20, weight="bold"))
app.imageFit_Title.grid(row=4, column=0, padx=20, pady=20,columnspan=2)

### Binning Section
app.FLIMAnalysis_Label = customtkinter.CTkLabel(app.Navigation_frame, text="Spatial Bin Size", compound="left", font=customtkinter.CTkFont(size=12, weight="bold"))
app.FLIMAnalysis_Label.grid(row=5, column = 0)

app.BinMethod_Label = customtkinter.CTkLabel(app.Navigation_frame, text="Bin Method", compound="left", font=customtkinter.CTkFont(size=12, weight="bold"))
app.BinMethod_Label.grid(row=6, column = 0)


app.inputBinningSize = customtkinter.CTkEntry(app.Navigation_frame, placeholder_text="3", width=30)
app.inputBinningSize.grid(row=5, column=1)

binningChoices = ['Column', 'Row', 'Square','Cell','None']
binngingVariable = customtkinter.StringVar(value="None")
wbin = customtkinter.CTkOptionMenu(app.Navigation_frame, values=binningChoices, variable=binngingVariable, command = Binning_Image_Way)
wbin.grid(row=6, column=1)

app.FlimAnalysis_BinButton = customtkinter.CTkButton(app.Navigation_frame,text="Spatial Bin", command= lambda: (Get_Binning(),Binning_Image(tpsfImageRT,binningSize,binningWay)),
                                                 corner_radius=0, height=40, border_spacing=10,
                                                 fg_color="transparent", text_color=("gray10", "gray90"),
                                                 hover_color=("gray70", "gray30"),
                                                 anchor="w")

app.FlimAnalysis_BinButton.grid(row=7, column=0,sticky="ew",columnspan=2)

### Fitting Section
app.FittingComponent_Label = customtkinter.CTkLabel(app.Navigation_frame, text="Fit Component", compound="left", font=customtkinter.CTkFont(size=12, weight="bold"))
app.FittingComponent_Label.grid(row=8, column = 0)
app.inputThresholdLabel = customtkinter.CTkLabel(app.Navigation_frame, text="Threshold", compound="left", font=customtkinter.CTkFont(size=12, weight="bold"))
app.inputThresholdLabel.grid(row=9, column=0)
app.FittingWayLabel = customtkinter.CTkLabel(app.Navigation_frame, text="Fit Method", compound="left", font=customtkinter.CTkFont(size=12, weight="bold"))
app.FittingWayLabel.grid(row=10, column=0)


componentChoices = ['Single', 'Bi']
componentVariable = customtkinter.StringVar(value="Single")
wFitComp = customtkinter.CTkOptionMenu(app.Navigation_frame, values = componentChoices, variable = componentVariable ,command = Fit_Component)
wFitComp.grid(row=8, column=1)

app.inputThreshold = customtkinter.CTkEntry(app.Navigation_frame, placeholder_text="10", width=50)
app.inputThreshold.grid(row=9, column=1)

fitMethods = ['MLE','Lsq','LsqN']
fitMethodsVariable = customtkinter.StringVar(value="MLE")
wFitMetd = customtkinter.CTkOptionMenu(app.Navigation_frame, values=fitMethods, variable=fitMethodsVariable,command = Fit_Method)
wFitMetd.grid(row=10, column=1)


app.fitCurveButton = customtkinter.CTkButton(app.Navigation_frame, corner_radius=0, height=40, border_spacing=10,
                                              text="FLIM Decay Curve Fit",
                                              fg_color="transparent", text_color=("gray10", "gray90"),
                                              hover_color=("gray70", "gray30"),
                                              anchor="w",command=lambda: (Get_Threshold(),FLIM_Fit()))
app.fitCurveButton.grid(row=11, column=0, sticky="ew")

app.fitProgress = customtkinter.CTkProgressBar(app.Navigation_frame,width = 2)
app.fitProgress.grid(row=11, column=1, sticky="ew",padx=20)
app.fitProgress.set(0)


# dispImageOptions = ['None']
dispImageVariable = customtkinter.StringVar(value="tm")
wDisOp = customtkinter.CTkOptionMenu(app.Navigation_frame, values=['None'], variable = dispImageVariable, command = Display_Image)
wDisOp.grid(row=12, column=0)


app.imageSaveButton = customtkinter.CTkButton(app.Navigation_frame, corner_radius=0, height=40, border_spacing=10,
                                              text="Save Lifetime Images",
                                              fg_color="transparent", text_color=("gray10", "gray90"),
                                              hover_color=("gray70", "gray30"),
                                              anchor="w", command = lambda: (Fit_Component,Image_Save(fitComponent,filePath,imgName)))
app.imageSaveButton.grid(row=12, column=1, sticky="ew")


### Cell Level Analysis
app.cellAnaly_Title = customtkinter.CTkLabel(app.Navigation_frame, text="Cell Level Analysis", compound="left", font=customtkinter.CTkFont(size=20, weight="bold"))
app.cellAnaly_Title.grid(row=13, column=0, padx=20, pady=20,columnspan=2)

app.calculateCellButton = customtkinter.CTkButton(app.Navigation_frame, corner_radius=0, height=40, border_spacing=10,
                                              text="Apply Mask",
                                              fg_color="transparent", text_color=("gray10", "gray90"),
                                              hover_color=("gray70", "gray30"),
                                              anchor="w", command=lambda: (Calculate_Life_Cell(a1Image, t1Image, t2Image, tmImage, tpsfImage, maskImage)))
app.calculateCellButton.grid(row=14, column=0, sticky="ew")


app.selectFileButton = customtkinter.CTkButton(app.Navigation_frame, corner_radius=0, height=40, border_spacing=10,
                                              text="Select Excel File",
                                              fg_color="transparent", text_color=("gray10", "gray90"),
                                              hover_color=("gray70", "gray30"),
                                              anchor="w", command=lambda: Select_Excel_File())
app.selectFileButton.grid(row=14, column=1, sticky="ew")

app.saveSheet = customtkinter.CTkEntry(app.Navigation_frame, placeholder_text="Sheet 1", width=30)
app.saveSheet.grid(row=15, column=0, sticky="ew")

app.saveExcelButton = customtkinter.CTkButton(app.Navigation_frame, corner_radius=0, height=40, border_spacing=10,
                                              text="Save Values",
                                              fg_color="transparent", text_color=("gray10", "gray90"),
                                              hover_color=("gray70", "gray30"),
                                              anchor="w", command=lambda: Save_Values_Excel(imgName))
app.saveExcelButton.grid(row=15, column=1, sticky="ew")




### Image Showing Home Frame
app.Image_frame = customtkinter.CTkFrame(app, corner_radius=0, fg_color="transparent",height=800,width=600)
app.Image_frame.grid(row=0, column=1, sticky="nsew")

app.Image_frame.grid_columnconfigure(0, weight=1)
app.Image_frame.grid_rowconfigure(0, weight=1)


FLIM_GAN_Model = tf.keras.models.load_model('FLIM_GAN_Test.h5',compile=False)

app.mainloop()






